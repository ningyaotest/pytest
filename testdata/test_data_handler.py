#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
@File  : test_data_handler.py
@Author: 姚宁
@Date  : 2021/5/23 22:33
@Desc  : 
'''
#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
@File  : test_data_handler.py
@Author: 姚宁
@Date  : 2021/4/21 21:08
@Desc  : 
'''
from openpyxl import load_workbook
import json
from openpyxl.workbook.workbook import Workbook, Worksheet
from openpyxl.cell.cell import Cell


def get_data_from_excel(file, sheet_name=None):
    wb = load_workbook(file)

    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]
    data=[]
    row = ws.max_row
    column = ws.max_column
    title = {}
    for i in range(1, column + 1):
        title[i] = ws.cell(row=1, column=i).value

    for j in range(2, row + 1):
        temp = {}
        for i in range(1, column + 1):
            values = ws.cell(row=j, column=i).value
            temp[title[i]] = values
        # temp[title[i]]=eval(temp[title[i]])
        data.append(temp)
    return data


if __name__ == '__main__':
    res = get_data_from_excel('testcase_data.xlsx', sheet_name='Alipay')
    s=(res[0]['service'])
    print(s,type(s))

